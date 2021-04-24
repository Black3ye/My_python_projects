import Tkinter
from Tkinter import Text
from pymsgbox import *
from tkcalendar import *
from openpyxl import *
from datetime import date

root = Tkinter.Tk()
root.title("S.Familia- Registro en excel")
root.geometry('700x400')

#Declaring Text Variables
incosto = Tkinter.DoubleVar()
inactividad = Tkinter.StringVar()
expense = Tkinter.StringVar()
gcosto = Tkinter.DoubleVar()
person = Tkinter.StringVar()

todate= date.today()
todate = todate.strftime("%Y")
fn='Registro'+todate+'.xlsx'
try:
	workbook = load_workbook(filename=fn)
except:
	fn =  'Registro'+todate+'.xlsx'
	workbook = Workbook(fn)
	ws1 = workbook.create_sheet("Ingresos")
	ws2 = workbook.create_sheet("Gastos")
	workbook.active = 0
	ws= workbook.active
	ws.append(["Actividad","Fecha","Cantidad","Notas","Total"])
	workbook.active = 1
	wsd2= workbook.active
	wsd2.append(["Gasto","Fecha","Cantidad","Recibido por","Notas","Total"])
	workbook.save(fn)
	

def clear():
	for obj in root.winfo_children():
            obj.destroy()
def ingresos():
	clear()
	def saving_in():
		notes= txtinnotas.get(1.0,Tkinter.END)
		workbook.active = 0
		wsheet = workbook.active
		rwn=len(wsheet['A']) + 1
		colin=['S','A','B','C','D','E']
		if (rwn==2):
			total= incosto.get()
		else:
			tr= str(rwn-1)
			total='=C'+str(rwn)+'+ E'+tr
		dl = [str(inactividad.get()),str(infecha.get_date()),str(incosto.get()),str(notes),str(total)]
		c=1
		d=0
		for i in range(len(colin)):
			if(i !=0):
				wsheet.cell(row=rwn, column=c).value= dl[d]
				c = c + 1
				d = d + 1
		workbook.save(fn)
		#https://pypi.org/project/PyMsgBox/
		confirm(title='S.Familia- Registro en excel',text='El registro de ingresos ha sido guardado',buttons=['Ok'])

	lblinactividad = Tkinter.Label(root,text='Actividad:').place(x=100, y=92)
	txtinactividad = Tkinter.Entry(root,textvariable= inactividad).place(x=170,y=90)
	lblinfecha = Tkinter.Label(root,text='Fecha:').place(x=360,y=92)
	#https://youtu.be/fqfy-3IoVvs
	infecha= Calendar(root,selectmode="day",year=2020,month=12,day=14)
	infecha.place(x=420,y=55)
	lblincosto = Tkinter.Label(root,text='Costo:').place(x=100,y=142)
	txtincosto = Tkinter.Entry(root,textvariable= incosto,width=10).place(x=165, y=140)
	lblinnotas = Tkinter.Label(root,text='Notas:').place(x=100,y=225)
	txtinnotas = Tkinter.Text(root,width=40,height=5)
	txtinnotas.place(x=150,y=225)
	btninsave = Tkinter.Button(root,text='Guardar',command=saving_in).place(x=610,y=365)
	btninback = Tkinter.Button(root,text='Back',command=menu).place(x=2,y=40)
	
def gastos():
	clear()
	def saving_out():
		notes= txtgnotas.get(1.0,Tkinter.END)
		workbook.active = 1
		wsheet = workbook.active
		rwn=len(wsheet['A']) + 1
		colin=['S','A','B','C','D','E','F']
		if (rwn==2):
			total=gcosto.get()
		else:
			tr=str(rwn-1)
			total= '=C'+str(rwn)+'+ F'+tr
		
		dlg = [str(expense.get()),str(gfecha.get_date()),str(gcosto.get()),str(person.get()),str(notes),str(total)]
		c=1
		d=0
		for i in range(len(colin)):
			if(i !=0):
				wsheet.cell(row=rwn, column=c).value= dlg[d]
				c = c + 1
				d = d + 1
		workbook.save(fn)
		#https://pypi.org/project/PyMsgBox/
		confirm(title='S.Familia- Registro en excel',text='El registro de gasto ha sido guardado',buttons=['Ok'])

	lblexpenses = Tkinter.Label(root,text='Gasto:').place(x=100, y=92)
	txtexpenses = Tkinter.Entry(root,textvariable= expense).place(x=150,y=90)
	lblgfecha = Tkinter.Label(root,text='Fecha:').place(x=360,y=92)
	#https://youtu.be/fqfy-3IoVvs
	gfecha= Calendar(root,selectmode="day",year=2020,month=12,day=14)
	gfecha.place(x=420,y=55)
	lblgcosto = Tkinter.Label(root,text='Costo:').place(x=100,y=137)
	txtgcosto = Tkinter.Entry(root,textvariable= gcosto,width=10).place(x=165, y=135)
	lblgpersona = Tkinter.Label(root,text='Comprados por:').place(x=100,y=180)
	txtgpersona = Tkinter.Entry(root,textvariable = person ).place(x=205,y=180)
	lblgnotas = Tkinter.Label(root,text='Notas:').place(x=100,y=225)
	txtgnotas = Tkinter.Text(root,width=40,height=5)
	txtgnotas.place(x=150,y=225)
	btngsave = Tkinter.Button(root,text='Guardar',command=saving_out).place(x=610,y=365)
	btngback = Tkinter.Button(root,text='Back',command=menu).place(x=2,y=40)
	

def menu():
	clear()
	lblwelcome = Tkinter.Label(root,text='Wecome Sofia').place(x=300,y=5)
	btnIngresos = Tkinter.Button(root,text='Ingresos',width=40,command=ingresos).place(x=0,y=40)
	btnGastos = Tkinter.Button(root,text='Gastos',width=40,command=gastos).place(x=350,y=40)
menu()
root.mainloop()
